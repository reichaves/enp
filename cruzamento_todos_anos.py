'''
Script para cruzamento estruturado de dados de trabalhadores resgatados sob condições 
análogas à escravidão no Brasil, correlacionando informações de três bases de dados governamentais 
e independentes (Planilha Manual de Autos, Perfil do Seguro-Desemprego e Radar SIT)

Autor: Reinaldo Chaves (reichaves@gmail.com)
https://github.com/reichaves

  ### 1. Pré-processamento e Padronização de Dados

  • Normalização de Documentos: As funções cruzamento_todos_anos.py e cruzamento_todos_anos.py removem caracteres
especiais ( . ,  - ,  /
  ) e preenchem com zeros à esquerda (11 caracteres para CPFs e 14 para CNPJs) para corrigir truncamentos de
  importação do Excel.
  • Padronização Textual: A função cruzamento_todos_anos.py remove acentos (normalização unicode  NFKD ), converte
caracteres
  para maiúsculas e remove espaços duplos ou marginais.
  • Tratamento de Chaves Vazias: Códigos de operação vazios ou nulos são convertidos de forma limpa, evitando a
  colisão e cruzamentos falsos por inconsistências de formatação.

  ### 2. Vinculação Trabalhador × Perfil (Seguro-Desemprego)

  • Fusão Hierárquica: O script tenta primeiro realizar o cruzamento por nome exato. Para os nomes remanescentes,
  utiliza a função cruzamento_todos_anos.py baseada na distância Levenshtein ( fuzz.ratio  da biblioteca  fuzzywuzzy )
com
  limite de similaridade ajustado em 90.
  • Resolução de Homônimos por Proximidade Temporal: Para homônimos correspondentes ao mesmo nome do trabalhador,
  calcula-se a diferença absoluta de dias entre a  Data Resgate  (Seguro-Desemprego) e a data de afastamento extraída
  do auto manual. O pipeline preserva apenas o registro de Seguro-Desemprego com menor intervalo de dias.
  • Validação Antihomônimos: A função  validar_sobrenomes  descarta correspondências caso os sobrenomes dos
  trabalhadores (excluindo preposições) não apresentem nenhuma palavra compartilhada.
  • Sinalização de Divergência Temporal: O script gera a coluna  flag_data_divergente  marcando como  True  os casos
  cuja diferença absoluta de dias de resgate excede 365 dias.

  ### 3. Vinculação Operação × Radar SIT

  O script unifica o cadastro de fiscalizações do Radar SIT e realiza junções (joins) baseadas em três níveis de
  prioridade sobre o ano fiscal e o empregador:

   Nível de Prioridade   │ Método de Cruzamento                     │ Colunas Envolvidas
  ───────────────────────┼──────────────────────────────────────────┼────────────────────────────────────────────────
   Prioridade 3 (Máxima) │ Correspondência Exata por Documento      │  cpf_cnpj_empregador_str  ↔  CNPJ-CEI-CPF_str
   Prioridade 2          │ Correspondência Exata por                │  estabelecimento_inspecionado_limpo  ↔
                         │ Estabelecimento                          │ Estabelecimento Inspecionado_Limpo
   Prioridade 1 (Mínima) │ Correspondência de Nomes Administrativos │ Razão Social ou Nome Fantasia ↔ Nome do
                         │                                          │ Estabelecimento ou Proprietário

  • Os registros sem correspondência no Radar SIT são mantidos na base consolidada (junção externa parcial) para
  impedir perda de registros de trabalhadores resgatados.

  ### 4. Auditoria de Estabelecimentos e Prazos

  • Classificação de Conformidade: A função cruzamento_todos_anos.py classifica a confiabilidade do cruzamento do
empregador em
  quatro status:
      •  VALIDADO_POR_CNPJ : CNPJs idênticos em ambas as bases.
      •  AUSENTE : Ausência de nomes de estabelecimentos para validação.
      •  COMPATIVEL : Sobreposição de termos textuais (overlap) de estabelecimentos $\ge 35\%$.
      •  DIVERGENTE : Sobreposição inferior a $35\%$.
  • Auditoria de Prazos: A coluna  checagem  avalia a proximidade de dias do resgate:
      • Até 90 dias: em branco.
      • De 91 a 365 dias: sinalizado como  verificar .
      • Acima de 365 dias: sinalizado como  divergente (>365d) .


  ### 5. Exportação Formatada

  O resultado final é exportado no formato Excel para o subdiretório  com o sufixo
  todos_anos_maio_2026_v11 . As colunas recebem preenchimento de cores estilizadas via  openpyxl :

  • Amarelo ( #FFF2CC ): Atributos originários dos Autos Manuais.
  • Verde ( #E2EFDA ): Dados de perfil do Seguro-Desemprego.
  • Azul ( #DDEBF7 ): Metadados provenientes do Radar SIT.

'''

import os
import pandas as pd
import numpy as np
import re
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.graph_objects as go

from unicodedata import normalize
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configuração visual dos gráficos
sns.set_theme(style="whitegrid")

# Definição dos Caminhos
path_dados = r"E:\Code\enp\dados_2024_2025"
path_resultados = r"E:\Code\enp\scripts\resultados_scripts"

# Sufixo atualizado para a versão 11
SUFIXO_SAIDA = "todos_anos_maio_2026_v11"

os.makedirs(path_resultados, exist_ok=True)

## 3. Funções de Padronização e Limpeza
def limpar_documento(doc):
    if pd.isna(doc): return ''
    doc_str = str(doc).strip()
    doc_str = re.sub(r'[\.\-\/]', '', doc_str)
    return doc_str

def padronizar_zeros_esquerda(doc, tipo='cpf'):
    if not doc: return ''
    if tipo == 'cpf': return doc.zfill(11)
    elif tipo == 'cnpj': return doc.zfill(14)
    return doc

def padronizar_coluna_texto(coluna):
    coluna = coluna.fillna('')
    coluna = coluna.str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')
    coluna = coluna.str.upper().str.strip()
    coluna = coluna.apply(lambda x: re.sub(r'\s+', ' ', x))
    return coluna

def remove_leading_zeros(num_str):
    if pd.isna(num_str) or str(num_str).strip() in ('', 'nan', 'NaN'): return ''
    val = str(num_str).strip()
    if val == '0': return '0'
    val = val.lstrip('0')
    return val if val else '0'

## 4. Carregamento das Bases de Dados
print("Carregando bases de dados...")

# Base 1: Seguro Desemprego
arquivo_perfil = os.path.join(path_dados, "ENP - Perfil do resgatado no Brasil.xlsx")
perfil_trabalhadores = pd.read_excel(arquivo_perfil, sheet_name='Dados', dtype='str')
cols_verde = perfil_trabalhadores.columns.tolist() + ['cpf_requerente_limpo', 'cpf_cnpj_empregador', 'Nome Requerente_Limpo']

perfil_trabalhadores["cpf_requerente_limpo"] = perfil_trabalhadores["CPF Requerente"].apply(limpar_documento).apply(lambda x: padronizar_zeros_esquerda(x, 'cpf'))
perfil_trabalhadores["cpf_cnpj_empregador"] = perfil_trabalhadores["Num Insc Empregador (CEI/CNPJ)"].apply(limpar_documento).apply(lambda x: padronizar_zeros_esquerda(x, 'cnpj'))
perfil_trabalhadores['Nome Requerente_Limpo'] = padronizar_coluna_texto(perfil_trabalhadores['Nome Requerente'])


# Base 2: Radar SIT
arquivo_local = os.path.join(path_dados, "radarsit_mai22.xlsx")
trabalhadores_local = pd.read_excel(arquivo_local, sheet_name='Planilha1', dtype='str')
cols_azul = trabalhadores_local.columns.tolist() + ['cpf_cnpj_limpo', 'Estabelecimento Inspecionado_Limpo', 'Proprierário_Limpo', 'CNPJ-CEI-CPF_str']

trabalhadores_local["Ano"] = trabalhadores_local["Ano"].astype(str).str.strip()
trabalhadores_local["Operação"] = trabalhadores_local["Operação"].apply(remove_leading_zeros)
trabalhadores_local["cpf_cnpj_limpo"] = trabalhadores_local["CNPJ-CEI-CPF"].apply(limpar_documento).apply(lambda x: padronizar_zeros_esquerda(x, 'cnpj'))
trabalhadores_local['Estabelecimento Inspecionado_Limpo'] = padronizar_coluna_texto(trabalhadores_local['Estabelecimento Inspecionado'])
trabalhadores_local['Proprierário_Limpo'] = padronizar_coluna_texto(trabalhadores_local['Proprierário'])


# Base 3: Autos Manuais
arquivo_autos = os.path.join(path_dados, "trabalhadores_mai22.xlsx")
trabalhadores_manual = pd.read_excel(arquivo_autos, sheet_name='Planilha1', dtype='str')
cols_amarelo = trabalhadores_manual.columns.tolist() + ['ano relatorio', 'operacao', 'nome_trabalhador_limpo', 'estabelecimento_inspecionado_limpo', 'cpf_cnpj_empregador_str', 'Razão Social Empregador_Limpa', 'Nome Fantasia Empregador_Limpo']

trabalhadores_manual['ano relatorio'] = trabalhadores_manual['ano relatorio'].astype(str).str.strip().replace(['nan', 'NaN', ''], np.nan)
trabalhadores_manual['operacao'] = trabalhadores_manual['operacao'].astype(str).str.strip().replace(['nan', 'NaN'], '')
trabalhadores_manual['nome_trabalhador_limpo'] = padronizar_coluna_texto(trabalhadores_manual['nome trabalhador'])
trabalhadores_manual['estabelecimento_inspecionado_limpo'] = padronizar_coluna_texto(trabalhadores_manual['estabelecimento inspecionado'])

## 5. Processamento dos Cruzamentos
# ETAPA 1
print("Realizando cruzamento por nome exato...")
merged_exato = pd.merge(trabalhadores_manual, perfil_trabalhadores, left_on='nome_trabalhador_limpo', right_on='Nome Requerente_Limpo', how='inner')
manuais_restantes = trabalhadores_manual[~trabalhadores_manual['nome_trabalhador_limpo'].isin(merged_exato['nome_trabalhador_limpo'])]

def get_best_match(name, candidates, cutoff=90):
    if not name: return None
    match = process.extractOne(name, candidates, scorer=fuzz.ratio, score_cutoff=cutoff)
    return match[0] if match else None

print(f"Realizando cruzamento Fuzzy para {len(manuais_restantes)} registros...")
candidatos_perfil = perfil_trabalhadores['Nome Requerente_Limpo'].tolist()
manuais_restantes_copy = manuais_restantes.copy()
manuais_restantes_copy['match_fuzzy'] = manuais_restantes_copy['nome_trabalhador_limpo'].apply(lambda x: get_best_match(x, candidatos_perfil))
merged_fuzzy = pd.merge(manuais_restantes_copy.dropna(subset=['match_fuzzy']), perfil_trabalhadores, left_on='match_fuzzy', right_on='Nome Requerente_Limpo', how='inner')
merged_trabalhadores = pd.concat([merged_exato, merged_fuzzy]).drop_duplicates()

# Filtragem por data
def parse_date_safe(val):
    if pd.isna(val) or str(val).strip() in ('', 'nan', 'None', 'NaT'): return pd.NaT
    try: return pd.to_datetime(str(val).strip()[:10], format='%Y-%m-%d', errors='coerce')
    except Exception: return pd.NaT

chave_manual = ['nome_trabalhador_limpo', 'ano relatorio', 'operacao']

if 'data afastamento/demissao' in merged_trabalhadores.columns and 'Data Resgate' in merged_trabalhadores.columns:
    merged_trabalhadores = merged_trabalhadores.copy()
    d_manual = merged_trabalhadores['data afastamento/demissao'].apply(parse_date_safe)
    d_sd     = merged_trabalhadores['Data Resgate'].apply(parse_date_safe)
    merged_trabalhadores['diferenca_dias_resgate'] = (d_sd - d_manual).abs().dt.days

    has_date    = d_manual.notna()
    with_date   = merged_trabalhadores[has_date].copy().sort_values('diferenca_dias_resgate', na_position='last').drop_duplicates(subset=chave_manual, keep='first')
    without_date = merged_trabalhadores[~has_date].copy()
    
    merged_trabalhadores = pd.concat([with_date, without_date], ignore_index=True)
    merged_trabalhadores['flag_data_divergente'] = (merged_trabalhadores['diferenca_dias_resgate'] > 365).fillna(False)

# TRAVA ABSOLUTA ETAPA 1
merged_trabalhadores = merged_trabalhadores.drop_duplicates(subset=chave_manual, keep='first').copy()

# SANITIZAÇÃO
print("Aplicando filtros...")
mask_data_div = merged_trabalhadores['flag_data_divergente'] == True
def validar_sobrenomes(row):
    n_man = str(row.get('nome_trabalhador_limpo', '')).lower().split()
    n_sd  = str(row.get('Nome Requerente_Limpo', '')).lower().split()
    if not n_man or not n_sd: return True
    prep = {'de', 'da', 'do', 'dos', 'das', 'e', 'a', 'o', 'van', 'von'}
    sob_man = set(n_man[1:]) - prep
    sob_sd  = set(n_sd[1:]) - prep
    if not sob_man or not sob_sd: return False
    return len(sob_man & sob_sd) == 0

mask_nome_div = merged_trabalhadores.apply(validar_sobrenomes, axis=1)
merged_trabalhadores = merged_trabalhadores[~(mask_data_div | mask_nome_div)].copy()

# ETAPA 2
print("Iniciando Etapa 2: Radar SIT...")
col_base_radar = ['Ano', 'Operação', 'CNPJ-CEI-CPF']
trabalhadores_local = trabalhadores_local.drop_duplicates(subset=col_base_radar, keep='first').copy()

merged_trabalhadores['Razão Social Empregador_Limpa'] = padronizar_coluna_texto(merged_trabalhadores['Razão Social Empregador'])
merged_trabalhadores['Nome Fantasia Empregador_Limpo'] = padronizar_coluna_texto(merged_trabalhadores['Nome Fantasia Empregador'])
merged_trabalhadores['cpf_cnpj_empregador_str'] = merged_trabalhadores['cpf_cnpj_empregador'].astype(str).str.strip().str.replace(r'[.\-/]', '', regex=True)
trabalhadores_local['CNPJ-CEI-CPF_str'] = trabalhadores_local['CNPJ-CEI-CPF'].astype(str).str.strip().str.replace(r'[.\-/]', '', regex=True)

colunas_chave = ['cpf_cnpj_empregador_str', 'CNPJ-CEI-CPF_str', 'estabelecimento_inspecionado_limpo', 'Estabelecimento Inspecionado_Limpo', 'Razão Social Empregador_Limpa', 'Proprierário_Limpo', 'Nome Fantasia Empregador_Limpo']
for df_alvo in [merged_trabalhadores, trabalhadores_local]:
    for col in colunas_chave:
        if col in df_alvo.columns:
            df_alvo[col] = df_alvo[col].replace(['', 'NAN', 'NaN', 'nan'], np.nan)

c_cnpj = pd.merge(merged_trabalhadores.dropna(subset=['cpf_cnpj_empregador_str']), trabalhadores_local.dropna(subset=['CNPJ-CEI-CPF_str']), left_on=['cpf_cnpj_empregador_str', 'ano relatorio'], right_on=['CNPJ-CEI-CPF_str', 'Ano'], how='inner')
c_cnpj['match_priority'] = 3
c_cnpj['match_tipo'] = 'CNPJ_E_ANO_EXATO'

c_est = pd.merge(merged_trabalhadores.dropna(subset=['estabelecimento_inspecionado_limpo']), trabalhadores_local.dropna(subset=['Estabelecimento Inspecionado_Limpo']), left_on=['estabelecimento_inspecionado_limpo', 'ano relatorio'], right_on=['Estabelecimento Inspecionado_Limpo', 'Ano'], how='inner')
c_est['match_priority'] = 2
c_est['match_tipo'] = 'ESTAB_E_ANO_EXATO'

c_raz = pd.merge(merged_trabalhadores.dropna(subset=['Razão Social Empregador_Limpa']), trabalhadores_local.dropna(subset=['Estabelecimento Inspecionado_Limpo']), left_on=['Razão Social Empregador_Limpa', 'ano relatorio'], right_on=['Estabelecimento Inspecionado_Limpo', 'Ano'], how='inner')
c_fant = pd.merge(merged_trabalhadores.dropna(subset=['Nome Fantasia Empregador_Limpo']), trabalhadores_local.dropna(subset=['Proprierário_Limpo']), left_on=['Nome Fantasia Empregador_Limpo', 'ano relatorio'], right_on=['Proprierário_Limpo', 'Ano'], how='inner')
c_prop = pd.merge(merged_trabalhadores.dropna(subset=['estabelecimento_inspecionado_limpo']), trabalhadores_local.dropna(subset=['Proprierário_Limpo']), left_on=['estabelecimento_inspecionado_limpo', 'ano relatorio'], right_on=['Proprierário_Limpo', 'Ano'], how='inner')

for df in [c_raz, c_fant, c_prop]:
    df['match_priority'] = 1
    df['match_tipo'] = 'NOME_EMPRESARIAL_E_ANO'

todos_matches = pd.concat([c_cnpj, c_est, c_raz, c_fant, c_prop], ignore_index=True)
cols_drop = [c for c in todos_matches.columns if c.endswith('_x') or c.endswith('_y')]
todos_matches = todos_matches.drop(columns=cols_drop, errors='ignore')

todos_matches['match_op_exata'] = (todos_matches['operacao'] == todos_matches['Operação']) & (todos_matches['operacao'] != '')
todos_matches = todos_matches.sort_values(by=['match_op_exata', 'match_priority'], ascending=[False, False])

key_cols_dedup = ['Nome Requerente_Limpo', 'ano relatorio', 'operacao']
best_matches = todos_matches.drop_duplicates(subset=key_cols_dedup, keep='first').copy()

keys_in_best = best_matches.set_index(key_cols_dedup).index
keys_all = merged_trabalhadores.set_index(key_cols_dedup).index
unmatched_mask = ~keys_all.isin(keys_in_best)
unmatched_trabalhadores = merged_trabalhadores[unmatched_mask].copy()

final_df = pd.concat([best_matches, unmatched_trabalhadores], ignore_index=True)

def valida_estab(row):
    if pd.isna(row.get('Ano')): return 'SEM_CORRESPONDENCIA_SIT' 
    
    # 1º AVALIA O CNPJ PRIMEIRO (Evita que o 'AUSENTE' roube o status do CNPJ validado)
    c1 = str(row.get('cpf_cnpj_empregador_str', '')).strip()
    c2 = str(row.get('CNPJ-CEI-CPF_str', '')).strip()
    if c1 and c2 and c1 == c2 and c1 != 'nan': return 'VALIDADO_POR_CNPJ'

    # 2º SE O CNPJ FALHOU OU NÃO EXISTE, AVALIA OS NOMES
    man = str(row.get('estabelecimento_inspecionado_limpo', '')).lower()
    rad = str(row.get('Estabelecimento Inspecionado_Limpo', '')).lower()
    
    if not man or not rad or man == 'nan' or rad == 'nan': return 'AUSENTE'

    prep = {'da','de','do','dos','das','e','s/a','s.a.','ltda','fazenda','faz','agro','me','eia','-','.'}
    w_man = set(man.split()) - prep
    w_rad = set(rad.split()) - prep
    if not w_man or not w_rad: return 'VALIDACAO_MANUAL'
    overlap = len(w_man & w_rad) / max(len(w_man), len(w_rad))
    return 'COMPATIVEL' if overlap >= 0.35 else 'DIVERGENTE'

final_df['flag_estabelecimento'] = final_df.apply(valida_estab, axis=1)

def regra_coluna_g(diff):
    if pd.isna(diff): return ''
    if float(diff) <= 90: return ''
    if 90 < float(diff) <= 365: return 'verificar'
    return 'divergente (>365d)'

final_df['checagem'] = final_df['diferenca_dias_resgate'].apply(regra_coluna_g)

print(f"Cruzamento concluído! Linhas únicas finais: {len(final_df)}")

## 6. Exportação de Resultados
caminho_exportacao = os.path.join(path_resultados, f"cruzamento_final_{SUFIXO_SAIDA}.xlsx")
final_df = final_df.loc[:, ~final_df.columns.duplicated()]
final_df.to_excel(caminho_exportacao, sheet_name='Resultados', index=False)

## 7. Formatação Visual (openpyxl)
print("Aplicando padrão de cores no Excel...")
wb = load_workbook(caminho_exportacao)
ws = wb['Resultados']

# Tons pastel para preservar a legibilidade
fill_amarelo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
fill_verde = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
fill_azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

for col_idx, col_name in enumerate(final_df.columns, 1):
    current_fill = None
    if col_name in cols_amarelo:
        current_fill = fill_amarelo
    elif col_name in cols_verde:
        current_fill = fill_verde
    elif col_name in cols_azul:
        current_fill = fill_azul

    if current_fill:
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).fill = current_fill

wb.save(caminho_exportacao)
print(f"Exportação finalizada com sucesso: {caminho_exportacao}")